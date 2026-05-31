from io import BytesIO
from datetime import date
from django.http import HttpResponse
from rest_framework.decorators import api_view
from django.db.models import Q
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    PageBreak
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.lib.colors import HexColor
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.pdfmetrics import registerFont

FONT_DIR = "C:/Windows/Fonts"
registerFont(TTFont('TimesNewRoman', f'{FONT_DIR}/times.ttf'))
registerFont(TTFont('TimesNewRoman-Bold', f'{FONT_DIR}/timesbd.ttf'))
registerFont(TTFont('TimesNewRoman-Italic', f'{FONT_DIR}/timesi.ttf'))
registerFont(TTFont('TimesNewRoman-BoldItalic', f'{FONT_DIR}/timesbi.ttf'))

from projects.models import Project, Site, Floor
from quantities.models import Element, MaterialQuantity
from scheduling.models import Activity

PRIMARY = HexColor('#1B3A5C')
SECONDARY = HexColor('#2C5F8A')
ACCENT = HexColor('#3498DB')
LIGHT_BG = HexColor('#F0F4F8')
CRITICAL_BG = HexColor('#FFE0E0')
DARK_TEXT = HexColor('#2C3E50')
MEDIUM_TEXT = HexColor('#7F8C8D')


def _header_footer(canvas, doc):
    canvas.saveState()
    w, h = A4
    pn = getattr(doc, 'project_name', 'Construction Project')
    rt = getattr(doc, 'report_title', 'Report')
    rn = getattr(doc, 'report_number', '')
    rd = getattr(doc, 'report_date', date.today().strftime('%Y-%m-%d'))
    pb = getattr(doc, 'prepared_by', '')

    # Header bar
    canvas.setFillColor(PRIMARY)
    canvas.rect(0, h - 20*mm, w, 20*mm, fill=1, stroke=0)

    # Header text
    canvas.setFillColor(colors.white)
    canvas.setFont('TimesNewRoman-Bold', 10)
    canvas.drawString(15*mm, h - 14*mm, pn[:60])
    canvas.setFont('TimesNewRoman', 8)
    canvas.drawRightString(w - 15*mm, h - 14*mm, f'{rt} | {rn}')

    # Footer bar
    canvas.setFillColor(PRIMARY)
    canvas.rect(0, 0, w, 12*mm, fill=1, stroke=0)

    # Footer text
    canvas.setFillColor(colors.white)
    canvas.setFont('TimesNewRoman', 7)
    canvas.drawString(15*mm, 4*mm, f'Issued: {rd} | Prepared: {pb}')
    canvas.drawRightString(w - 15*mm, 4*mm, f'Page {doc.page}')

    canvas.restoreState()


def _build_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        'ReportTitle', fontName='TimesNewRoman-Bold', fontSize=18, leading=22,
        alignment=TA_CENTER, spaceAfter=2*mm, textColor=PRIMARY
    ))
    styles.add(ParagraphStyle(
        'ReportSubtitle', fontName='TimesNewRoman', fontSize=9, leading=12,
        alignment=TA_CENTER, spaceAfter=1*mm, textColor=MEDIUM_TEXT
    ))
    styles.add(ParagraphStyle(
        'SectionTitle', fontName='TimesNewRoman-Bold', fontSize=11, leading=14,
        alignment=TA_LEFT, spaceBefore=4*mm, spaceAfter=2*mm,
        textColor=SECONDARY, borderWidth=0
    ))
    styles.add(ParagraphStyle(
        'SubSectionTitle', fontName='TimesNewRoman-Bold', fontSize=9, leading=12,
        alignment=TA_LEFT, spaceBefore=2*mm, spaceAfter=1*mm, textColor=DARK_TEXT
    ))
    styles.add(ParagraphStyle(
        'ReportBody', fontName='TimesNewRoman', fontSize=8, leading=11,
        alignment=TA_LEFT, spaceAfter=1*mm, textColor=DARK_TEXT
    ))
    styles.add(ParagraphStyle(
        'SmallText', fontName='TimesNewRoman', fontSize=7, leading=9,
        alignment=TA_LEFT, textColor=MEDIUM_TEXT
    ))
    styles.add(ParagraphStyle(
        'TableCell', fontName='TimesNewRoman', fontSize=8, leading=11,
        alignment=TA_LEFT, textColor=DARK_TEXT
    ))
    styles.add(ParagraphStyle(
        'TableCellCenter', fontName='TimesNewRoman', fontSize=8, leading=11,
        alignment=TA_CENTER, textColor=DARK_TEXT
    ))
    styles.add(ParagraphStyle(
        'TableCellRight', fontName='TimesNewRoman', fontSize=8, leading=11,
        alignment=TA_RIGHT, textColor=DARK_TEXT
    ))
    styles.add(ParagraphStyle(
        'TableHeader', fontName='TimesNewRoman-Bold', fontSize=8, leading=11,
        alignment=TA_CENTER, textColor=colors.white
    ))
    styles.add(ParagraphStyle(
        'TableHeaderLeft', fontName='TimesNewRoman-Bold', fontSize=8, leading=11,
        alignment=TA_LEFT, textColor=colors.white
    ))
    return styles


def _make_table_style(header_rows=1, critical_rows=None):
    cmds = [
        ('GRID', (0, 0), (-1, -1), 0.4, HexColor('#CCCCCC')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]
    if header_rows > 0:
        cmds += [
            ('BACKGROUND', (0, 0), (-1, header_rows - 1), PRIMARY),
            ('TEXTCOLOR', (0, 0), (-1, header_rows - 1), colors.white),
            ('FONTNAME', (0, 0), (-1, header_rows - 1), 'TimesNewRoman-Bold'),
            ('FONTSIZE', (0, 0), (-1, header_rows - 1), 9),
            ('ALIGN', (0, 0), (-1, header_rows - 1), 'CENTER'),
            ('MINIMUM HEIGHT', (0, 0), (-1, header_rows - 1), 6*mm),
        ]
    cmds += [
        ('FONTNAME', (0, header_rows), (-1, -1), 'TimesNewRoman'),
        ('FONTSIZE', (0, header_rows), (-1, -1), 9),
        ('MINIMUM HEIGHT', (0, header_rows), (-1, -1), 5*mm),
        ('ALIGN', (0, header_rows), (-1, -1), 'LEFT'),
        ('ROWBACKGROUNDS', (0, header_rows), (-1, -1), [colors.white, LIGHT_BG]),
    ]
    if critical_rows:
        for row in critical_rows:
            if row >= header_rows:
                cmds.append(('BACKGROUND', (0, row), (-1, row), HexColor('#FFF0F0')))
    return TableStyle(cmds)


def _info_block(styles, items):
    """Create a key-value info table."""
    data = []
    for key, val in items:
        data.append([
            Paragraph(f'<b>{key}</b>', styles['SmallText']),
            Paragraph(str(val), styles['SmallText'])
        ])
    t = Table(data, colWidths=[50*mm, 130*mm])
    t.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
        ('LINEBELOW', (0, 0), (-1, -1), 0.3, HexColor('#E0E0E0')),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]))
    return t


def _build_report_base(project, report_title, report_number, styles, prepared_by=''):
    elements = []

    # Title block
    elements.append(Paragraph(report_title, styles['ReportTitle']))
    if project:
        elements.append(Paragraph(f'{project.name} — {project.location}', styles['ReportSubtitle']))
    elements.append(Spacer(1, 3*mm))

    # Meta info block
    today = date.today().strftime('%Y-%m-%d')
    meta = [
        ('Report No.', report_number),
        ('Issue Date', today),
        ('Project', project.name if project else '-'),
        ('Project Type', project.project_type if project else '-'),
        ('Prepared By', prepared_by or 'Engineering Department'),
    ]
    elements.append(_info_block(styles, meta))
    elements.append(Spacer(1, 4*mm))

    return elements


# ──────────────────────────────────────────────
# 1. QUANTITY TAKEOFF PDF
# ──────────────────────────────────────────────

@api_view(['GET'])
def quantity_takeoff_pdf(request, project_id):
    try:
        project = Project.objects.get(id=project_id)
    except Project.DoesNotExist:
        return HttpResponse('Project not found', status=404)

    sites = Site.objects.filter(project=project).prefetch_related('floors__elements__quantities')
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            topMargin=25*mm, bottomMargin=17*mm,
                            leftMargin=15*mm, rightMargin=15*mm)
    styles = _build_styles()
    elements = []
    total_materials = {}

    # Meta info
    report_no = f'QTO-{project_id:04d}'
    elements += _build_report_base(project, 'QUANTITY TAKEOFF REPORT', report_no, styles,
                                   prepared_by=request.query_params.get('prepared_by', ''))

    # Per-site/floor breakdown
    for site in sites:
        elements.append(Paragraph(f'Site: {site.name}', styles['SectionTitle']))
        floors = Floor.objects.filter(site=site).order_by('floor_number')
        for floor in floors:
            els = Element.objects.filter(floor=floor).prefetch_related('quantities')
            els_list = list(els)
            if not els_list:
                continue
            elements.append(Paragraph(f'Floor: {floor.name} (Level {floor.floor_number})', styles['SubSectionTitle']))

            data = [[
                Paragraph('Element', styles['TableHeaderLeft']),
                Paragraph('Type', styles['TableHeader']),
                Paragraph('Count', styles['TableHeader']),
                Paragraph('Material', styles['TableHeader']),
                Paragraph('Quantity', styles['TableHeader']),
                Paragraph('Unit', styles['TableHeader']),
                Paragraph('Specification', styles['TableHeader']),
            ]]
            for el in els_list:
                row_start = len(data)
                for mq in el.quantities.all():
                    data.append([
                        Paragraph(el.name or '-', styles['TableCell']),
                        Paragraph(el.element_type.replace('_', ' ').title(), styles['TableCell']),
                        Paragraph(str(el.count), styles['TableCellCenter']),
                        Paragraph(mq.material_type.replace('_', ' ').title(), styles['TableCell']),
                        Paragraph(f'{mq.value:,.2f}', styles['TableCellRight']),
                        Paragraph(mq.unit, styles['TableCellCenter']),
                        Paragraph(mq.spec_reference, styles['SmallText']),
                    ])
                    key = mq.material_type
                    total_materials[key] = total_materials.get(key, 0) + mq.value
                # If no quantities, show a placeholder row
                if len(data) == row_start:
                    data.append([
                        Paragraph(el.name or el.element_type, styles['TableCell']),
                        Paragraph(el.element_type.replace('_', ' ').title(), styles['TableCell']),
                        Paragraph(str(el.count), styles['TableCellCenter']),
                        Paragraph('-', styles['TableCellCenter']),
                        Paragraph('-', styles['TableCellRight']),
                        Paragraph('-', styles['TableCellCenter']),
                        Paragraph('-', styles['SmallText']),
                    ])

            if len(data) > 1:
                t = Table(data, colWidths=[50*mm, 30*mm, 12*mm, 32*mm, 22*mm, 14*mm, 36*mm])
                t.setStyle(_make_table_style())
                elements.append(t)
                elements.append(Spacer(1, 2*mm))

    # Infrastructure section
    infra_els = Element.objects.filter(project=project).prefetch_related('quantities')
    if infra_els.exists():
        elements.append(Spacer(1, 4*mm))
        elements.append(Paragraph('INFRASTRUCTURE WORKS', styles['SectionTitle']))
        data = [[
            Paragraph('Element', styles['TableHeaderLeft']),
            Paragraph('Type', styles['TableHeader']),
            Paragraph('Count', styles['TableHeader']),
            Paragraph('Material', styles['TableHeader']),
            Paragraph('Quantity', styles['TableHeader']),
            Paragraph('Unit', styles['TableHeader']),
            Paragraph('Specification', styles['TableHeader']),
        ]]
        for el in infra_els:
            row_start = len(data)
            for mq in el.quantities.all():
                data.append([
                    Paragraph(el.name or '-', styles['TableCell']),
                    Paragraph(el.element_type.replace('_', ' ').title(), styles['TableCell']),
                    Paragraph(str(el.count), styles['TableCellCenter']),
                    Paragraph(mq.material_type.replace('_', ' ').title(), styles['TableCell']),
                    Paragraph(f'{mq.value:,.2f}', styles['TableCellRight']),
                    Paragraph(mq.unit, styles['TableCellCenter']),
                    Paragraph(mq.spec_reference, styles['SmallText']),
                ])
                key = mq.material_type
                total_materials[key] = total_materials.get(key, 0) + mq.value
            if len(data) == row_start:
                data.append([
                    Paragraph(el.name or el.element_type, styles['TableCell']),
                    Paragraph(el.element_type.replace('_', ' ').title(), styles['TableCell']),
                    Paragraph(str(el.count), styles['TableCellCenter']),
                    Paragraph('-', styles['TableCellCenter']),
                    Paragraph('-', styles['TableCellRight']),
                    Paragraph('-', styles['TableCellCenter']),
                    Paragraph('-', styles['SmallText']),
                ])
        if len(data) > 1:
            t = Table(data, colWidths=[50*mm, 30*mm, 12*mm, 32*mm, 22*mm, 14*mm, 36*mm])
            t.setStyle(_make_table_style())
            elements.append(t)

    # Summary section
    elements.append(Spacer(1, 5*mm))
    elements.append(Paragraph('MATERIAL SUMMARY', styles['SectionTitle']))
    summary_data = [[
        Paragraph('Material', styles['TableHeader']),
        Paragraph('Total Quantity', styles['TableHeader']),
        Paragraph('Unit', styles['TableHeader']),
    ]]
    for mat, qty in sorted(total_materials.items()):
        summary_data.append([
            Paragraph(mat.replace('_', ' ').title(), styles['TableCell']),
            Paragraph(f'{qty:,.2f}', styles['TableCellRight']),
            Paragraph('-', styles['TableCellCenter']),
        ])
    t2 = Table(summary_data, colWidths=[100*mm, 60*mm, 30*mm])
    t2.setStyle(_make_table_style())
    elements.append(t2)

    doc.project_name = project.name
    doc.report_title = 'QUANTITY TAKEOFF'
    doc.report_number = report_no
    doc.report_date = date.today().strftime('%Y-%m-%d')
    doc.prepared_by = request.query_params.get('prepared_by', 'Engineering Department')
    doc.build(elements, onFirstPage=_header_footer, onLaterPages=_header_footer)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf',
                        headers={'Content-Disposition': f'attachment; filename="QTO_{project_id:04d}.pdf"'})


# ──────────────────────────────────────────────
# 2. BENDING SCHEDULE PDF
# ──────────────────────────────────────────────

@api_view(['GET'])
def bending_schedule_pdf(request, project_id):
    try:
        project = Project.objects.get(id=project_id)
    except Project.DoesNotExist:
        return HttpResponse('Project not found', status=404)

    sites = Site.objects.filter(project=project)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            topMargin=25*mm, bottomMargin=17*mm,
                            leftMargin=15*mm, rightMargin=15*mm)
    styles = _build_styles()

    report_no = f'BBS-{project_id:04d}'
    elements = _build_report_base(project, 'REINFORCING STEEL BENDING SCHEDULE', report_no, styles,
                                  prepared_by=request.query_params.get('prepared_by', ''))

    has_data = False
    for site in sites:
        floors = Floor.objects.filter(site=site).order_by('floor_number')
        for floor in floors:
            els = Element.objects.filter(floor=floor)
            for el in els:
                rebar_qty = MaterialQuantity.objects.filter(element=el, material_type='rebar').first()
                if not rebar_qty or rebar_qty.value == 0:
                    continue
                has_data = True
                elements.append(Paragraph(
                    f'{site.name} / {floor.name} / {el.name or el.element_type} &mdash; '
                    f'Total Rebar: {rebar_qty.value:,.2f} {rebar_qty.unit}',
                    styles['SubSectionTitle']
                ))

                from quantities.calculators.registry import CALCULATOR_MAP
                calc_cls = CALCULATOR_MAP.get(el.element_type)
                bars = []
                if calc_cls:
                    try:
                        calc = calc_cls(el.dimensions, el.count)
                        bars = calc.extra_results().get('bending_schedule', [])
                    except Exception:
                        pass

                data = [[
                    Paragraph('Bar Dia (mm)', styles['TableHeader']),
                    Paragraph('US Size', styles['TableHeader']),
                    Paragraph('Count', styles['TableHeader']),
                    Paragraph('Length (m)', styles['TableHeader']),
                    Paragraph('Total Length (m)', styles['TableHeader']),
                    Paragraph('Weight (kg)', styles['TableHeader']),
                    Paragraph('Shape Code', styles['TableHeader']),
                    Paragraph('Shape Description', styles['TableHeader']),
                ]]
                if bars:
                    for b in bars:
                        data.append([
                            Paragraph(str(b.get('dia', '')), styles['TableCellCenter']),
                            Paragraph(b.get('us_size', '-'), styles['TableCellCenter']),
                            Paragraph(str(b.get('count', '')), styles['TableCellCenter']),
                            Paragraph(f'{b.get("length_m", 0):,.2f}', styles['TableCellRight']),
                            Paragraph(f'{b.get("total_length_m", 0):,.2f}', styles['TableCellRight']),
                            Paragraph(f'{b.get("weight_kg", 0):,.2f}', styles['TableCellRight']),
                            Paragraph(b.get('shape_code', '-'), styles['TableCellCenter']),
                            Paragraph(b.get('shape_desc', '-'), styles['SmallText']),
                        ])
                else:
                    data.append([
                        Paragraph('-', styles['TableCellCenter']),
                        Paragraph('-', styles['TableCellCenter']),
                        Paragraph('-', styles['TableCellCenter']),
                        Paragraph('-', styles['TableCellRight']),
                        Paragraph('-', styles['TableCellRight']),
                        Paragraph('-', styles['TableCellRight']),
                        Paragraph('-', styles['TableCellCenter']),
                        Paragraph('No bending details available', styles['SmallText']),
                    ])

                t = Table(data, colWidths=[24*mm, 18*mm, 16*mm, 24*mm, 30*mm, 24*mm, 20*mm, 54*mm])
                t.setStyle(_make_table_style())
                elements.append(t)
                elements.append(Spacer(1, 2*mm))

    if not has_data:
        elements.append(Paragraph('No reinforcing steel items found in this project.', styles['ReportBody']))

    doc.project_name = project.name
    doc.report_title = 'BENDING SCHEDULE'
    doc.report_number = report_no
    doc.report_date = date.today().strftime('%Y-%m-%d')
    doc.prepared_by = request.query_params.get('prepared_by', 'Engineering Department')
    doc.build(elements, onFirstPage=_header_footer, onLaterPages=_header_footer)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf',
                        headers={'Content-Disposition': f'attachment; filename="BBS_{project_id:04d}.pdf"'})


# ──────────────────────────────────────────────
# 3. SCHEDULE / CPM REPORT PDF
# ──────────────────────────────────────────────

@api_view(['GET'])
def schedule_pdf(request, project_id):
    try:
        project = Project.objects.get(id=project_id)
    except Project.DoesNotExist:
        return HttpResponse('Project not found', status=404)

    activities = list(Activity.objects.filter(project=project).order_by('order'))
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            topMargin=25*mm, bottomMargin=17*mm,
                            leftMargin=15*mm, rightMargin=15*mm)
    styles = _build_styles()

    report_no = f'CPM-{project_id:04d}'
    elements = _build_report_base(project, 'PROJECT SCHEDULE — CPM REPORT', report_no, styles,
                                  prepared_by=request.query_params.get('prepared_by', ''))

    if activities:
        project_dur = max((a.early_finish for a in activities), default=0)
        critical_names = [a.name for a in activities if a.is_critical]

        elements.append(Paragraph('SCHEDULE SUMMARY', styles['SectionTitle']))
        summary_items = [
            ('Total Duration', f'{project_dur} days'),
            ('Total Activities', str(len(activities))),
            ('Critical Activities', str(sum(1 for a in activities if a.is_critical))),
            ('Critical Path', ' → '.join(critical_names) if critical_names else '-'),
        ]
        elements.append(_info_block(styles, summary_items))
        elements.append(Spacer(1, 4*mm))

        elements.append(Paragraph('ACTIVITY SCHEDULE', styles['SectionTitle']))
        data = [[
            Paragraph('#', styles['TableHeader']),
            Paragraph('Activity', styles['TableHeader']),
            Paragraph('Type', styles['TableHeader']),
            Paragraph('Dur.', styles['TableHeader']),
            Paragraph('ES', styles['TableHeader']),
            Paragraph('EF', styles['TableHeader']),
            Paragraph('LS', styles['TableHeader']),
            Paragraph('LF', styles['TableHeader']),
            Paragraph('Float', styles['TableHeader']),
            Paragraph('Critical', styles['TableHeader']),
        ]]
        critical_indices = []
        for idx, a in enumerate(activities):
            data.append([
                Paragraph(str(a.order or idx + 1), styles['TableCellCenter']),
                Paragraph(a.name, styles['TableCell']),
                Paragraph(a.activity_type.replace('_', ' ').title(), styles['TableCell']),
                Paragraph(str(a.duration_days), styles['TableCellCenter']),
                Paragraph(str(a.early_start), styles['TableCellCenter']),
                Paragraph(str(a.early_finish), styles['TableCellCenter']),
                Paragraph(str(a.late_start), styles['TableCellCenter']),
                Paragraph(str(a.late_finish), styles['TableCellCenter']),
                Paragraph(str(a.total_float), styles['TableCellCenter']),
                Paragraph('Yes' if a.is_critical else '', styles['TableCellCenter']),
            ])
            if a.is_critical:
                critical_indices.append(idx + 1)

        t = Table(data, colWidths=[12*mm, 60*mm, 28*mm, 14*mm, 14*mm, 14*mm, 14*mm, 14*mm, 14*mm, 16*mm])
        t.setStyle(_make_table_style(critical_rows=critical_indices))
        elements.append(t)
    else:
        elements.append(Paragraph('No activities have been created yet.', styles['ReportBody']))

    doc.project_name = project.name
    doc.report_title = 'SCHEDULE CPM'
    doc.report_number = report_no
    doc.report_date = date.today().strftime('%Y-%m-%d')
    doc.prepared_by = request.query_params.get('prepared_by', 'Engineering Department')
    doc.build(elements, onFirstPage=_header_footer, onLaterPages=_header_footer)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf',
                        headers={'Content-Disposition': f'attachment; filename="CPM_{project_id:04d}.pdf"'})


# ──────────────────────────────────────────────
# 4. BILL OF QUANTITIES EXCEL
# ──────────────────────────────────────────────

@api_view(['GET'])
def boq_excel(request, project_id):
    try:
        project = Project.objects.get(id=project_id)
    except Project.DoesNotExist:
        return HttpResponse('Project not found', status=404)

    sites = Site.objects.filter(project=project)
    today = date.today()
    report_no = f'BOQ-{project_id:04d}'

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'BOQ'
    ws.sheet_properties.tabColor = '1B3A5C'

    # Colors
    primary_fill = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
    secondary_fill = PatternFill(start_color='2C5F8A', end_color='2C5F8A', fill_type='solid')
    light_fill = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid')
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    title_font = Font(name='Calibri', bold=True, color='1B3A5C', size=14)
    subtitle_font = Font(name='Calibri', color='7F8C8D', size=9)
    data_font = Font(name='Calibri', size=9)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    # Title section
    ws.cell(row=1, column=1, value='BILL OF QUANTITIES').font = title_font
    ws.merge_cells('A1:H1')
    ws.cell(row=2, column=1, value=f'{project.name} | {project.location}').font = subtitle_font
    ws.merge_cells('A2:H2')
    ws.cell(row=3, column=1, value=f'Report: {report_no} | Date: {today}').font = subtitle_font
    ws.merge_cells('A3:H3')

    # Headers
    headers = ['Site', 'Floor', 'Element', 'Type', 'Material', 'Quantity', 'Unit', 'Specification']
    header_row = 5
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=ci, value=h)
        c.font = header_font
        c.fill = primary_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border

    # Data
    row = header_row + 1
    for site in sites:
        floors = Floor.objects.filter(site=site).order_by('floor_number')
        for floor in floors:
            els = Element.objects.filter(floor=floor).prefetch_related('quantities')
            for el in els:
                has_qty = False
                for mq in el.quantities.all():
                    ws.cell(row=row, column=1, value=site.name).font = data_font
                    ws.cell(row=row, column=1).border = thin_border
                    ws.cell(row=row, column=2, value=floor.name).font = data_font
                    ws.cell(row=row, column=2).border = thin_border
                    ws.cell(row=row, column=3, value=el.name or '-').font = data_font
                    ws.cell(row=row, column=3).border = thin_border
                    ws.cell(row=row, column=4, value=el.element_type.replace('_', ' ').title()).font = data_font
                    ws.cell(row=row, column=4).border = thin_border
                    ws.cell(row=row, column=5, value=mq.material_type.replace('_', ' ').title()).font = data_font
                    ws.cell(row=row, column=5).border = thin_border
                    c = ws.cell(row=row, column=6, value=mq.value)
                    c.font = data_font
                    c.number_format = '#,##0.00'
                    c.alignment = Alignment(horizontal='right')
                    c.border = thin_border
                    ws.cell(row=row, column=7, value=mq.unit).font = data_font
                    ws.cell(row=row, column=7).border = thin_border
                    ws.cell(row=row, column=8, value=mq.spec_reference).font = data_font
                    ws.cell(row=row, column=8).border = thin_border
                    row += 1
                    has_qty = True
                if not has_qty:
                    ws.cell(row=row, column=1, value=site.name).font = data_font
                    ws.cell(row=row, column=1).border = thin_border
                    ws.cell(row=row, column=2, value=floor.name).font = data_font
                    ws.cell(row=row, column=2).border = thin_border
                    ws.cell(row=row, column=3, value=el.name or el.element_type).font = data_font
                    ws.cell(row=row, column=3).border = thin_border
                    ws.cell(row=row, column=4, value=el.element_type.replace('_', ' ').title()).font = data_font
                    ws.cell(row=row, column=4).border = thin_border
                    for ci in range(5, 9):
                        ws.cell(row=row, column=ci, value='-').font = data_font
                        ws.cell(row=row, column=ci).border = thin_border
                    row += 1

    # Infrastructure items
    infra_els = Element.objects.filter(project=project).prefetch_related('quantities')
    for el in infra_els:
        has_qty = False
        for mq in el.quantities.all():
            ws.cell(row=row, column=1, value='Infrastructure').font = data_font
            ws.cell(row=row, column=1).border = thin_border
            ws.cell(row=row, column=2, value='-').font = data_font
            ws.cell(row=row, column=2).border = thin_border
            ws.cell(row=row, column=3, value=el.name or '-').font = data_font
            ws.cell(row=row, column=3).border = thin_border
            ws.cell(row=row, column=4, value=el.element_type.replace('_', ' ').title()).font = data_font
            ws.cell(row=row, column=4).border = thin_border
            ws.cell(row=row, column=5, value=mq.material_type.replace('_', ' ').title()).font = data_font
            ws.cell(row=row, column=5).border = thin_border
            c = ws.cell(row=row, column=6, value=mq.value)
            c.font = data_font
            c.number_format = '#,##0.00'
            c.alignment = Alignment(horizontal='right')
            c.border = thin_border
            ws.cell(row=row, column=7, value=mq.unit).font = data_font
            ws.cell(row=row, column=7).border = thin_border
            ws.cell(row=row, column=8, value=mq.spec_reference).font = data_font
            ws.cell(row=row, column=8).border = thin_border
            row += 1
            has_qty = True
        if not has_qty:
            ws.cell(row=row, column=1, value='Infrastructure').font = data_font
            ws.cell(row=row, column=1).border = thin_border
            ws.cell(row=row, column=2, value='-').font = data_font
            ws.cell(row=row, column=2).border = thin_border
            ws.cell(row=row, column=3, value=el.name or el.element_type).font = data_font
            ws.cell(row=row, column=3).border = thin_border
            ws.cell(row=row, column=4, value=el.element_type.replace('_', ' ').title()).font = data_font
            ws.cell(row=row, column=4).border = thin_border
            for ci in range(5, 9):
                ws.cell(row=row, column=ci, value='-').font = data_font
                ws.cell(row=row, column=ci).border = thin_border
            row += 1

    # Alternate row coloring
    for r in range(header_row + 1, row):
        if (r - header_row) % 2 == 0:
            for ci in range(1, 9):
                ws.cell(row=r, column=ci).fill = light_fill

    # Column widths
    widths = [14, 14, 18, 16, 18, 14, 8, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Footer info
    row += 1
    ws.cell(row=row, column=1, value='Prepared By: Engineering Department').font = subtitle_font
    ws.merge_cells(f'A{row}:H{row}')

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition': f'attachment; filename="BOQ_{project_id:04d}.xlsx"'})


# ──────────────────────────────────────────────
# 5. BILL OF QUANTITIES WITH PRICES EXCEL
# ──────────────────────────────────────────────

def _get_or_generate_estimate(project):
    """Return latest estimate or auto-generate one."""
    from costing.models import CostEstimate, CostItem, MaterialPrice
    from quantities.models import Element, MaterialQuantity

    estimate = CostEstimate.objects.filter(project=project).order_by('-id').first()
    if estimate:
        return estimate

    prices = {p.material_type: p for p in MaterialPrice.objects.all()}
    estimate = CostEstimate.objects.create(project=project, markup_percent=15)
    elements = Element.objects.filter(Q(floor__site__project=project) | Q(project=project)).prefetch_related('quantities')
    for el in elements:
        for mq in el.quantities.all():
            price = prices.get(mq.material_type)
            unit_price = price.unit_price if price else 0
            if mq.value > 0:
                CostItem.objects.create(
                    estimate=estimate, element=el,
                    material_type=mq.material_type,
                    description=f'{el.name or el.element_type} - {mq.material_type}',
                    quantity=mq.value, unit=mq.unit, unit_price=unit_price,
                )
    return estimate


@api_view(['GET'])
def boq_prices_excel(request, project_id):
    try:
        project = Project.objects.get(id=project_id)
    except Project.DoesNotExist:
        return HttpResponse('Project not found', status=404)

    estimate = _get_or_generate_estimate(project)

    today = date.today()
    report_no = f'BOQ-P-{project_id:04d}'

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'BOQ Prices'
    ws.sheet_properties.tabColor = '1B3A5C'

    primary_fill = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
    accent_fill = PatternFill(start_color='2C5F8A', end_color='2C5F8A', fill_type='solid')
    light_fill = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid')
    green_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    title_font = Font(name='Calibri', bold=True, color='1B3A5C', size=14)
    subtitle_font = Font(name='Calibri', color='7F8C8D', size=9)
    data_font = Font(name='Calibri', size=9)
    bold_font = Font(name='Calibri', bold=True, size=10, color='1B3A5C')
    total_font = Font(name='Calibri', bold=True, size=11, color='1B3A5C')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    # Title
    ws.cell(row=1, column=1, value='BILL OF QUANTITIES WITH PRICES').font = title_font
    ws.merge_cells('A1:H1')
    ws.cell(row=2, column=1, value=f'{project.name} | {project.location}').font = subtitle_font
    ws.merge_cells('A2:H2')
    ws.cell(row=3, column=1, value=f'Report: {report_no} | Date: {today}').font = subtitle_font
    ws.merge_cells('A3:H3')

    # Headers
    headers = ['Description', 'Material', 'Quantity', 'Unit', 'Unit Price (EGP)', 'Total (EGP)']
    header_row = 5
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=ci, value=h)
        c.font = header_font
        c.fill = primary_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border

    # Data
    items = list(estimate.items.all().order_by('material_type'))
    row = header_row + 1
    total_material = 0
    for item in items:
        ws.cell(row=row, column=1, value=item.description).font = data_font
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2, value=item.material_type.replace('_', ' ').title()).font = data_font
        ws.cell(row=row, column=2).border = thin_border
        c = ws.cell(row=row, column=3, value=item.quantity)
        c.font = data_font; c.number_format = '#,##0.00'; c.alignment = Alignment(horizontal='right'); c.border = thin_border
        ws.cell(row=row, column=4, value=item.unit).font = data_font
        ws.cell(row=row, column=4).border = thin_border
        c = ws.cell(row=row, column=5, value=item.unit_price)
        c.font = data_font; c.number_format = '#,##0.00'; c.alignment = Alignment(horizontal='right'); c.border = thin_border
        c = ws.cell(row=row, column=6, value=item.total_cost)
        c.font = data_font; c.number_format = '#,##0.00'; c.alignment = Alignment(horizontal='right'); c.border = thin_border
        total_material += item.total_cost
        if (row - header_row) % 2 == 0:
            for ci in range(1, 7):
                ws.cell(row=row, column=ci).fill = light_fill
        row += 1

    # Summary section
    row += 1
    ws.cell(row=row, column=1, value='SUMMARY').font = bold_font
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    # Total Material
    ws.cell(row=row, column=1, value='Total Material Cost:').font = bold_font
    ws.cell(row=row, column=1).border = thin_border
    ws.merge_cells(f'A{row}:E{row}')
    for ci in range(1, 6):
        ws.cell(row=row, column=ci).fill = green_fill
        ws.cell(row=row, column=ci).border = thin_border
    c = ws.cell(row=row, column=6, value=total_material)
    c.font = total_font; c.number_format = '#,##0.00'; c.fill = green_fill; c.border = thin_border
    row += 1

    # Markup
    markup_val = total_material * estimate.markup_percent / 100
    ws.cell(row=row, column=1, value=f'Overhead & Profit ({estimate.markup_percent}%):').font = bold_font
    ws.cell(row=row, column=1).border = thin_border
    ws.merge_cells(f'A{row}:E{row}')
    for ci in range(1, 6):
        ws.cell(row=row, column=ci).fill = green_fill
        ws.cell(row=row, column=ci).border = thin_border
    c = ws.cell(row=row, column=6, value=markup_val)
    c.font = total_font; c.number_format = '#,##0.00'; c.fill = green_fill; c.border = thin_border
    row += 1

    # Grand Total
    grand_total = total_material + markup_val
    ws.cell(row=row, column=1, value='GRAND TOTAL (incl. markup):').font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    ws.cell(row=row, column=1).border = thin_border
    ws.merge_cells(f'A{row}:E{row}')
    for ci in range(1, 6):
        ws.cell(row=row, column=ci).fill = primary_fill
        ws.cell(row=row, column=ci).border = thin_border
    c = ws.cell(row=row, column=6, value=grand_total)
    c.font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    c.fill = primary_fill; c.number_format = '#,##0.00'; c.border = thin_border

    # Column widths
    widths = [40, 18, 14, 10, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition': f'attachment; filename="BOQ_P_{project_id:04d}.xlsx"'})


# ──────────────────────────────────────────────
# 6. BILL OF QUANTITIES WITH PRICES PDF
# ──────────────────────────────────────────────

@api_view(['GET'])
def boq_prices_pdf(request, project_id):
    try:
        project = Project.objects.get(id=project_id)
    except Project.DoesNotExist:
        return HttpResponse('Project not found', status=404)

    estimate = _get_or_generate_estimate(project)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            topMargin=25*mm, bottomMargin=17*mm,
                            leftMargin=15*mm, rightMargin=15*mm)
    styles = _build_styles()
    report_no = f'BOQ-P-{project_id:04d}'
    elements = _build_report_base(project, 'BILL OF QUANTITIES WITH PRICES', report_no, styles,
                                  prepared_by=request.query_params.get('prepared_by', ''))

    items = list(estimate.items.all().order_by('material_type'))
    if items:
        data = [[
            Paragraph('Description', styles['TableHeader']),
            Paragraph('Material', styles['TableHeader']),
            Paragraph('Qty', styles['TableHeader']),
            Paragraph('Unit', styles['TableHeader']),
            Paragraph('Unit Price', styles['TableHeader']),
            Paragraph('Total (EGP)', styles['TableHeader']),
        ]]
        for item in items:
            data.append([
                Paragraph(item.description, styles['TableCell']),
                Paragraph(item.material_type.replace('_', ' ').title(), styles['TableCell']),
                Paragraph(f'{item.quantity:,.2f}', styles['TableCellRight']),
                Paragraph(item.unit, styles['TableCellCenter']),
                Paragraph(f'{item.unit_price:,.2f}', styles['TableCellRight']),
                Paragraph(f'{item.total_cost:,.2f}', styles['TableCellRight']),
            ])

        t = Table(data, colWidths=[70*mm, 30*mm, 22*mm, 16*mm, 24*mm, 26*mm])
        t.setStyle(_make_table_style())
        elements.append(t)

        # Summary
        elements.append(Spacer(1, 5*mm))
        total_material = sum(item.total_cost for item in items)
        markup_val = total_material * estimate.markup_percent / 100
        grand_total = total_material + markup_val

        summary_data = [
            [Paragraph('', styles['TableCell']),
             Paragraph('', styles['TableCell']),
             Paragraph('', styles['TableCell']),
             Paragraph('', styles['TableCell']),
             Paragraph('', styles['TableCell']),
             Paragraph('', styles['TableCell'])],
            [Paragraph('<b>Total Material Cost</b>', styles['TableCellRight']),
             Paragraph('', styles['TableCell']),
             Paragraph('', styles['TableCell']),
             Paragraph('', styles['TableCell']),
             Paragraph('', styles['TableCell']),
             Paragraph(f'{total_material:,.2f} EGP', styles['TableCellRight'])],
        ]
        st = Table(summary_data, colWidths=[70*mm, 30*mm, 22*mm, 16*mm, 24*mm, 26*mm])
        st.setStyle(TableStyle([
            ('LINEABOVE', (0, 0), (-1, 0), 1, PRIMARY),
            ('LINEBELOW', (0, -1), (-1, -1), 0.5, HexColor('#CCCCCC')),
            ('FONTNAME', (0, 0), (-1, -1), 'TimesNewRoman'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))
        elements.append(st)

        # Grand total box
        elements.append(Spacer(1, 3*mm))
        gt_data = [[
            Paragraph(f'<b>GRAND TOTAL (incl. {estimate.markup_percent}% markup): {grand_total:,.2f} EGP</b>', styles['ReportTitle'])
        ]]
        gt = Table(gt_data, colWidths=[188*mm])
        gt.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), PRIMARY),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, -1), 4*mm),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4*mm),
        ]))
        elements.append(gt)
    else:
        elements.append(Paragraph('No items found in estimate.', styles['ReportBody']))

    doc.project_name = project.name
    doc.report_title = 'BOQ WITH PRICES'
    doc.report_number = report_no
    doc.report_date = date.today().strftime('%Y-%m-%d')
    doc.prepared_by = request.query_params.get('prepared_by', 'Engineering Department')
    doc.build(elements, onFirstPage=_header_footer, onLaterPages=_header_footer)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf',
                        headers={'Content-Disposition': f'attachment; filename="BOQ_P_{project_id:04d}.pdf"'})
